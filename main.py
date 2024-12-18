from tabula import read_pdf
import pandas as pd
import numpy as np
import os
from tqdm import tqdm
from openpyxl import load_workbook


def is_currency(value):
    if pd.isna(value) or value == '':
        return False
    try:
        float(str(value).replace(',', ''))
    except ValueError:
        return False
    else:
        return True


def clean_numeric_columns(dataframe, columns):

    for column in columns:
        dataframe[column] = dataframe[column].str.replace(',', '')
        dataframe[column] = pd.to_numeric(dataframe[column], errors='coerce')
        dataframe[column] = dataframe[column].astype('float')

    return dataframe


def union_source(dataframes):

    dfs = []
    for temp_df in dataframes:

        # Split DB into new column
        temp_df[['amount', 'type']] = temp_df[4].str.extract(r'([\d,]+(?:\.\d+)?)\s*(DB|CR)?')
        temp_df = temp_df.drop(temp_df.columns[4], axis=1)

        if(len(temp_df.columns) == 7):
            # Name column and reorder
            temp_df.columns = ['date', 'desc', 'detail', 'branch', 'balance', 'amount', 'type']
            temp_df = temp_df[['date', 'desc', 'detail', 'branch', 'amount', 'type', 'balance']]

            dfs.append(temp_df)

    df = pd.concat(dfs, ignore_index=True)
    df = df.fillna(value=np.nan)
                
    return df


def insert_shifted_column(dataframe):

    # Add new columns with shifted values for comparison
    dataframe['prev_date'] = dataframe['date'].shift(1)
    dataframe['prev_desc'] = dataframe['desc'].shift(1)
    dataframe['prev_detail'] = dataframe['detail'].shift(1)
    dataframe['prev_branch'] = dataframe['branch'].shift(1)
    dataframe['prev_amount'] = dataframe['amount'].shift(1)
    dataframe['prev_transaction_type'] = dataframe['type'].shift(1)
    dataframe['prev_balance'] = dataframe['balance'].shift(1)

    dataframe = dataframe.fillna(value=np.nan)

    return dataframe


def extract_transactions(dataframe):

    transactions = []
    details = []
    descs = []
    temp = {}

    for index, row in dataframe.iterrows():

        if (row['desc'] == 'DR KOREKSI BUNGA') or (row['desc'] == 'BUNGA'):
            transaction = {
                "date": temp['date'],
                "desc": ' | '.join(descs) if descs else '',
                "detail": ' | '.join(details) if details else '',
                "branch": temp['branch'],
                "amount": temp['amount'],
                "transaction_type": temp['transaction_type'] if temp['transaction_type'] == 'DB' else 'CR',
                "balance": temp['balance']
            }
            transactions.append(transaction)
            break

        if (row['desc'] == 'SALDO AWAL'):
            continue

        # New Transaction
        if(not pd.isna(row['amount'])) and ((pd.isna(row['prev_amount'])) or row['amount'] != row['prev_amount']):
            # Save previous transaction
            if temp:
                transaction = {
                    "date": temp['date'],
                    "desc": ' | '.join(descs) if descs else '',
                    "detail": ' | '.join(details) if details else '',
                    "branch": temp['branch'],
                    "amount": temp['amount'],
                    "transaction_type": temp['transaction_type'] if temp['transaction_type'] == 'DB' else 'CR',
                    "balance": temp['balance']
                }
                transactions.append(transaction)
                details = []
                descs = []
                temp = {}

            temp = {
                'date': row['date'],
                'branch': row['branch'],
                'amount': row['amount'],
                'transaction_type': row['type'],
                'balance': row['balance']
            }

        if (not pd.isna(row['desc'])):
            descs.append(row['desc'])
        if (not pd.isna(row['detail'])):
            details.append(row['detail'])

    transaction_dataframe = pd.DataFrame(transactions)

    return transaction_dataframe


def calculate_balance(dataframe):

    dataframe['balance'] = init_balance
    # Iterate over rows
    for index, row in dataframe.iterrows():
        # If transaction type is 'DB', subtract amount from balance
        if row['transaction_type'] == 'DB':
            if index == 0:
                # For the first row, subtract amount from init_balance
                dataframe.at[index, 'balance'] -= row['amount']
            else:
                # For subsequent rows, subtract amount from the previous row's balance
                dataframe.at[index, 'balance'] = dataframe.at[index - 1, 'balance'] - row['amount']
        # If transaction type is 'CR', add amount to balance
        elif row['transaction_type'] == 'CR':
            if index == 0:
                # For the first row, add amount to init_balance
                dataframe.at[index, 'balance'] += row['amount']
            else:
                # For subsequent rows, add amount to the previous row's balance
                dataframe.at[index, 'balance'] = dataframe.at[index - 1, 'balance'] + row['amount']

    return dataframe


def save_to_excel(dataframe, output_filename):

    if os.path.isfile(output_filename):
        writer = pd.ExcelWriter(output_filename, engine="openpyxl", mode='a', if_sheet_exists='replace')
    else:
        writer = pd.ExcelWriter(output_filename, engine="openpyxl")

    dataframe.to_excel(writer, sheet_name=periode, index=False)

    workbook = writer.book
    worksheet = writer.sheets[periode]

    # Format column into currency IDR
    for cell in worksheet['E']:
        cell.number_format = '_-Rp* #,##0.00_-;[Red]-Rp* #,##0.00_-;_-Rp* "-"_-;_-@_-'
    for cell in worksheet['G']:
        cell.number_format = '_-Rp* #,##0.00_-;[Red]-Rp* #,##0.00_-;_-Rp* "-"_-;_-@_-'

    # Autofit column width
    for column_cells in worksheet.columns:
        max_length = max(len(str(cell.value)) for cell in column_cells)
        if str(column_cells[0].column) in ['5', '7']:
            worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 10
        else:
            worksheet.column_dimensions[column_cells[0].column_letter].width = max_length + 2

    writer.close()

    return


def get_year_month(sheet_name):
    year, month_name = sheet_name.split(' ')
    month_dict = {
        'JANUARI': 1, 'FEBRUARI': 2, 'MARET': 3, 'APRIL': 4,
        'MEI': 5, 'JUNI': 6, 'JULI': 7, 'AGUSTUS': 8,
        'SEPTEMBER': 9, 'OKTOBER': 10, 'NOVEMBER': 11, 'DESEMBER': 12
    }
    return int(year), month_dict[month_name]


def reorder_sheets(output_filename):

    wb = load_workbook(output_filename)
    sheet_names = wb.sheetnames
    sorted_sheets = sorted(sheet_names, key=get_year_month, reverse=True)
    wb._sheets.sort(key=lambda x: sorted_sheets.index(x.title))
    wb.save(output_filename)

    return

statements_folder = "statements"
pbar = tqdm(os.listdir(statements_folder))
for filename in pbar:
    file_path = os.path.join(statements_folder, filename)
    pbar.set_description("Processing %s" % filename)

    if os.path.isfile(file_path):

        # Get header information
        header_dataframe = read_pdf(file_path, area = (70, 315, 141, 548), pages='1', pandas_options={'header': None, 'dtype': str}, force_subprocess=True)[0]
        periode = header_dataframe.loc[header_dataframe[0] == 'PERIODE', 2].values[0]
        periode = ' '.join(reversed(periode.split()))
        no_rekening = header_dataframe.loc[header_dataframe[0] == 'NO. REKENING', 2].values[0]
        output_filename = f'{no_rekening}.xlsx'

        # y1, x1, y2, x2
        dataframes = read_pdf(file_path, area = (231, 25, 797, 577), columns=[86, 184, 300, 340, 467], pages='all', pandas_options={'header': None, 'dtype': str}, force_subprocess=True)

        init_balance = dataframes[0].loc[dataframes[0][1] == 'SALDO AWAL', 5].values[0]
        init_balance = float(init_balance.replace(',', ''))

        df = union_source(dataframes)
        df = clean_numeric_columns(df, ['amount', 'balance'])
        df = insert_shifted_column(df)

        transaction_dataframe = extract_transactions(df)
        transaction_dataframe = transaction_dataframe.drop('balance', axis=1)
        transaction_dataframe = calculate_balance(transaction_dataframe)

        save_to_excel(transaction_dataframe, output_filename)

reorder_sheets(output_filename)